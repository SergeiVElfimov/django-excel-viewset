from io import BytesIO

import openpyxl
from xlsxwriter import Workbook

from django_excel_viewset.excel_writer import XLSXCreator
from django_excel_viewset.excel_writer.creator import XLSXCell, XLSXRow, XLSXTable


class TestXLSXCreator:
    """Test the XLSXCreator."""

    @classmethod
    def setup_class(cls):
        cls.file = BytesIO()
        cls._workbook = Workbook(cls.file, {"in_memory": True})
        cls._worksheet = cls._workbook.add_worksheet(name="test")
        cls._xlsx_creator = XLSXCreator(
            workbook=cls._workbook,
            worksheet=cls._worksheet,
            datetime_cell_format={},
            date_cell_format={},
            percent_cell_format={},
            format_cell_border={},
            format_header_table={},
            table_label_format={},
        )
        cls._headers = ("Head 1", "Head 2")
        cls._data = [[1, 2], [3, 4]]

    def test_add_table(self):
        """Test the add_table method."""
        self._xlsx_creator.add_table(heading=self._headers, data=self._data, table_label="Test table")
        assert len(self._xlsx_creator.blocks) == 1
        assert isinstance(self._xlsx_creator.blocks[0], XLSXTable)

    def test_add_cell(self):
        """Test the add_cell method."""
        self._xlsx_creator.add_cell(
            text="test text", cell_format={"bold": True, "text_wrap": True, "font_size": 15}, row_padding=2
        )
        assert len(self._xlsx_creator.blocks) == 2
        assert isinstance(self._xlsx_creator.blocks[1], XLSXCell)

    def test_add_row(self):
        """Test the add_row method."""
        self._xlsx_creator.add_row(row=["test", 3], row_padding=1)
        assert len(self._xlsx_creator.blocks) == 3
        assert isinstance(self._xlsx_creator.blocks[2], XLSXRow)

    def test_make_excel(self):
        """Test the make_excel method."""
        self._xlsx_creator.make_excel()
        dataframe = openpyxl.load_workbook(filename=self.file)
        dataframe1 = dataframe.active
        res_data = [["test", None], [None, None], ["Head 1", "Head 2"], [1, 2], [3, 4]]
        for row in range(0, dataframe1.max_row):
            i = 0
            for col in dataframe1.iter_cols(1, dataframe1.max_column):
                assert col[row].value == res_data[row][i]
                i = i + 1
