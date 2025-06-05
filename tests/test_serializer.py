from io import BytesIO

import openpyxl
from rest_framework.serializers import CharField, IntegerField, Serializer

from django_excel_viewset import ExcelSerializerMixin


class TestExcelSerializerMixin:
    @classmethod
    def setup_class(cls):
        class ExampleSerializer(ExcelSerializerMixin, Serializer):
            text = CharField(help_text="Text")
            integer = IntegerField(help_text="Number")

        cls._serializer_class = ExampleSerializer

    def test_property(self):
        """Test excel property serializer."""
        serializer = self._serializer_class(data=[{"text": "test text", "integer": 1}], many=True)
        assert serializer.is_valid()
        assert serializer.excel_header == ("Text", "Number")
        assert serializer.excel_rows == [["test text", 1]]

    def test_excel_response(self):
        """Test excel_response method."""
        init_data = [{"text": "test text 1", "integer": 1}, {"text": "test text 2", "integer": 2}]
        serializer = self._serializer_class(data=init_data, many=True)
        assert serializer.is_valid()
        response = serializer.excel_response

        result = [list(serializer.excel_header)]
        for item in init_data:
            result.append(list(item.values()))

        workbook = openpyxl.load_workbook(filename=BytesIO(response.content))
        worksheet = workbook.active
        for row in range(0, worksheet.max_row):
            i = 0
            for col in worksheet.iter_cols(1, worksheet.max_column):
                assert col[row].value == result[row][i]
                i = i + 1
